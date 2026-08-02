import io
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from fletchscore.io.export.excel import exporter_classement_excel, exporter_classement_global_excel
from fletchscore.models import BAREME_FLINT_INDOOR, Competiteur, Epreuve, Score, Sexe, StatutScore
from fletchscore.scoring import classement_global, classement_par_categorie


def _competiteur(id_federal: str, sexe: Sexe) -> Competiteur:
    return Competiteur(
        id_federal=id_federal,
        nom=f"Nom-{id_federal}",
        prenom=f"Prenom-{id_federal}",
        code_club="77123",
        sexe=sexe,
        date_naissance=date(1995, 1, 1),
        code_style="BB-R",
    )


def _score(inscription_id: str, total: int) -> Score:
    return Score(
        id=f"s-{inscription_id}",
        inscription_id=inscription_id,
        total=total,
        statut=StatutScore.VALIDE,
    )


class TestExporterClassementExcel(unittest.TestCase):
    def setUp(self):
        entrees = [
            (_competiteur("FR-1", Sexe.M), _score("i1", 20)),
            (_competiteur("FR-2", Sexe.F), _score("i2", 19)),
        ]
        self.classement = classement_par_categorie(BAREME_FLINT_INDOOR, date(2026, 1, 1), entrees)

    def _charger(self, destination) -> list[list]:
        classeur = load_workbook(destination)
        feuille = classeur.active
        return [[cellule.value for cellule in ligne] for ligne in feuille.iter_rows()]

    def test_produit_un_fichier_valide_dans_un_flux_binaire(self):
        destination = io.BytesIO()
        exporter_classement_excel(self.classement, destination)
        destination.seek(0)
        lignes = self._charger(destination)
        self.assertGreater(len(lignes), 0)

    def test_deux_categories_donnent_deux_titres_en_gras(self):
        destination = io.BytesIO()
        exporter_classement_excel(self.classement, destination)
        destination.seek(0)
        classeur = load_workbook(destination)
        feuille = classeur.active

        titres_gras = [
            cellule.value
            for ligne in feuille.iter_rows()
            for cellule in ligne
            if cellule.column == 1
            and cellule.font
            and cellule.font.bold
            and cellule.value not in ("Rang",)
        ]
        self.assertEqual(set(titres_gras), {"AMBB-R", "AFBB-R"})

    def test_entetes_de_colonnes_presentes_par_categorie(self):
        destination = io.BytesIO()
        exporter_classement_excel(self.classement, destination)
        destination.seek(0)
        lignes = self._charger(destination)
        premieres_colonnes = [ligne[0] for ligne in lignes]
        self.assertEqual(premieres_colonnes.count("Rang"), 2)  # 1 par catégorie

    def test_valeurs_dune_ligne_competiteur(self):
        destination = io.BytesIO()
        exporter_classement_excel(self.classement, destination)
        destination.seek(0)
        lignes = self._charger(destination)
        ligne_fr1 = next(ligne for ligne in lignes if ligne[1] == "FR-1")
        self.assertEqual(ligne_fr1[0], 1)  # rang
        self.assertEqual(ligne_fr1[4], 20)  # total
        self.assertIsNone(ligne_fr1[5])  # pas de X sur ce barème (Flint)

    def test_classement_vide_produit_quand_meme_un_fichier(self):
        destination = io.BytesIO()
        exporter_classement_excel({}, destination)
        destination.seek(0)
        lignes = self._charger(destination)
        self.assertEqual(lignes[0][0], "Aucun compétiteur classé.")

    def test_ecriture_dans_un_fichier_reel(self):
        with tempfile.TemporaryDirectory() as dossier:
            chemin = str(Path(dossier) / "classement.xlsx")
            exporter_classement_excel(self.classement, chemin)
            self.assertTrue(Path(chemin).exists())
            self.assertGreater(Path(chemin).stat().st_size, 0)

    def test_titre_feuille_tronque_a_31_caracteres(self):
        destination = io.BytesIO()
        titre_trop_long = "Un titre de feuille vraiment beaucoup trop long pour Excel"
        exporter_classement_excel(self.classement, destination, titre_feuille=titre_trop_long)
        destination.seek(0)
        classeur = load_workbook(destination)
        self.assertLessEqual(len(classeur.active.title), 31)


class TestExporterClassementGlobalExcel(unittest.TestCase):
    def setUp(self):
        self.epreuve1 = Epreuve(
            id="epr-1", competition_id="comp-1", nom="Indoor", date=date(2026, 3, 14),
            bareme_id="ifaa-indoor",
        )
        self.epreuve2 = Epreuve(
            id="epr-2", competition_id="comp-1", nom="Flint", date=date(2026, 3, 15),
            bareme_id="flint-indoor",
        )
        competiteur = _competiteur("FR-1", Sexe.M)
        entrees = [(competiteur, {"epr-1": _score("i1", 260), "epr-2": _score("i2", 220)})]
        self.classement = classement_global(date(2026, 3, 14), ["epr-1", "epr-2"], entrees)

    def _charger(self, destination) -> list[list]:
        classeur = load_workbook(destination)
        feuille = classeur.active
        return [[cellule.value for cellule in ligne] for ligne in feuille.iter_rows()]

    def test_entete_contient_une_colonne_par_epreuve(self):
        destination = io.BytesIO()
        exporter_classement_global_excel(
            [self.epreuve1, self.epreuve2], self.classement, destination
        )
        destination.seek(0)
        lignes = self._charger(destination)
        # ligne 0 = titre catégorie, ligne 1 = en-têtes de colonnes
        self.assertIn("Indoor", lignes[1])
        self.assertIn("Flint", lignes[1])
        self.assertIn("Total", lignes[1])

    def test_valeurs_dune_ligne(self):
        destination = io.BytesIO()
        exporter_classement_global_excel(
            [self.epreuve1, self.epreuve2], self.classement, destination
        )
        destination.seek(0)
        lignes = self._charger(destination)
        ligne_fr1 = next(ligne for ligne in lignes if ligne[1] == "FR-1")
        # rang, id_federal, nom, prenom, epr1, epr2, total, x
        self.assertEqual(ligne_fr1[4], 260)
        self.assertEqual(ligne_fr1[5], 220)
        self.assertEqual(ligne_fr1[6], 480)

    def test_classement_vide_produit_quand_meme_un_fichier(self):
        destination = io.BytesIO()
        exporter_classement_global_excel([self.epreuve1, self.epreuve2], {}, destination)
        destination.seek(0)
        lignes = self._charger(destination)
        self.assertEqual(lignes[0][0], "Aucun compétiteur classé.")


if __name__ == "__main__":
    unittest.main()
