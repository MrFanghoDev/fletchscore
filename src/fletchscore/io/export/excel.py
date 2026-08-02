"""Export Excel du classement -- une feuille, groupée par catégorie.

Format destiné à la fédération (voir
docs/cahier-des-charges/modele-donnees.rst §6.2) -- le format exact
imposé ou non par la FFTL reste un point ouvert (voir docs/roadmap.md) ;
en attendant, ce tableau reprend les mêmes colonnes que l'export CSV.
"""

from __future__ import annotations

from typing import BinaryIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from fletchscore.models import Epreuve
from fletchscore.scoring import LigneClassement, LigneClassementGlobal

_ENTETES_COLONNES = ("Rang", "Id fédéral", "Nom", "Prénom", "Total", "X")
_LARGEURS_COLONNES = (6, 14, 18, 18, 8, 6)


def exporter_classement_excel(
    classement: dict[str, list[LigneClassement]],
    destination: str | BinaryIO,
    titre_feuille: str = "Classement",
) -> None:
    """Écrit le classement dans un classeur Excel (.xlsx), une section
    par catégorie (triées alphabétiquement) sur une seule feuille."""
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = titre_feuille[:31]  # limite Excel : 31 caractères max

    for index, largeur in enumerate(_LARGEURS_COLONNES, start=1):
        lettre = feuille.cell(row=1, column=index).column_letter
        feuille.column_dimensions[lettre].width = largeur

    ligne_courante = 1
    if not classement:
        feuille.cell(row=1, column=1, value="Aucun compétiteur classé.")
    else:
        for categorie in sorted(classement):
            ligne_courante = _ecrire_categorie(
                feuille, ligne_courante, categorie, classement[categorie]
            )

    _ecrire_sortie(classeur, destination)


def _ecrire_categorie(
    feuille: Worksheet,
    ligne: int,
    categorie: str,
    lignes_classement: list[LigneClassement],
) -> int:
    cellule_titre = feuille.cell(row=ligne, column=1, value=categorie)
    cellule_titre.font = Font(bold=True, size=13)
    ligne += 1

    for colonne, entete in enumerate(_ENTETES_COLONNES, start=1):
        cellule = feuille.cell(row=ligne, column=colonne, value=entete)
        cellule.font = Font(bold=True)
        cellule.alignment = Alignment(horizontal="center")
    ligne += 1

    for ligne_classement in lignes_classement:
        _ecrire_ligne_competiteur(feuille, ligne, ligne_classement)
        ligne += 1

    return ligne + 1  # une ligne vide avant la catégorie suivante


def _ecrire_ligne_competiteur(
    feuille: Worksheet, ligne: int, ligne_classement: LigneClassement
) -> None:
    competiteur = ligne_classement.competiteur
    valeurs = (
        ligne_classement.rang,
        competiteur.id_federal,
        competiteur.nom,
        competiteur.prenom,
        ligne_classement.total,
        ligne_classement.nombre_x if ligne_classement.nombre_x else None,
    )
    for colonne, valeur in enumerate(valeurs, start=1):
        feuille.cell(row=ligne, column=colonne, value=valeur)


def _ecrire_sortie(classeur: Workbook, destination: str | BinaryIO) -> None:
    classeur.save(destination)


def exporter_classement_global_excel(
    epreuves: list[Epreuve],
    classement: dict[str, list[LigneClassementGlobal]],
    destination: str | BinaryIO,
    titre_feuille: str = "Classement",
) -> None:
    """Écrit le classement cumulé d'une compétition -- une colonne par
    épreuve, plus une colonne total (voir
    ``io.export.csv.exporter_classement_global_csv`` pour le détail du
    format, même principe ici en Excel)."""
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = titre_feuille[:31]

    entetes = ("Rang", "Id fédéral", "Nom", "Prénom", *[e.nom for e in epreuves], "Total", "X")
    largeurs = (6, 14, 18, 18, *([12] * len(epreuves)), 8, 6)
    for index, largeur in enumerate(largeurs, start=1):
        lettre = feuille.cell(row=1, column=index).column_letter
        feuille.column_dimensions[lettre].width = largeur

    if not classement:
        feuille.cell(row=1, column=1, value="Aucun compétiteur classé.")
    else:
        ligne_courante = 1
        for categorie in sorted(classement):
            ligne_courante = _ecrire_categorie_globale(
                feuille, ligne_courante, categorie, classement[categorie], epreuves, entetes
            )

    _ecrire_sortie(classeur, destination)


def _ecrire_categorie_globale(
    feuille: Worksheet,
    ligne: int,
    categorie: str,
    lignes_classement: list[LigneClassementGlobal],
    epreuves: list[Epreuve],
    entetes: tuple[str, ...],
) -> int:
    cellule_titre = feuille.cell(row=ligne, column=1, value=categorie)
    cellule_titre.font = Font(bold=True, size=13)
    ligne += 1

    for colonne, entete in enumerate(entetes, start=1):
        cellule = feuille.cell(row=ligne, column=colonne, value=entete)
        cellule.font = Font(bold=True)
        cellule.alignment = Alignment(horizontal="center")
    ligne += 1

    for ligne_classement in lignes_classement:
        competiteur = ligne_classement.competiteur
        totaux = [ligne_classement.totaux_par_epreuve.get(e.id, 0) for e in epreuves]
        valeurs = (
            ligne_classement.rang,
            competiteur.id_federal,
            competiteur.nom,
            competiteur.prenom,
            *totaux,
            ligne_classement.total_global,
            ligne_classement.nombre_x_global if ligne_classement.nombre_x_global else None,
        )
        for colonne, valeur in enumerate(valeurs, start=1):
            feuille.cell(row=ligne, column=colonne, value=valeur)
        ligne += 1

    return ligne + 1
